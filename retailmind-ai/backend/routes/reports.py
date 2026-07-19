"""
RetailMind AI - Reports Routes
"""
import os
import io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import get_jwt_identity

from models import db, Product, Offer, ActivityLog
from utils import admin_required, log_activity

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


def get_report_data(period):
    """Get report data for the specified period."""
    now = datetime.utcnow()
    
    if period == 'daily':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == 'weekly':
        start = now - timedelta(days=7)
    elif period == 'monthly':
        start = now - timedelta(days=30)
    else:
        start = now - timedelta(days=1)
    
    # Products created in period
    new_products = Product.query.filter(Product.created_at >= start).all()
    
    # Offers created in period
    new_offers = Offer.query.filter(Offer.created_at >= start).all()
    
    # Active offers
    active_offers = Offer.query.filter(
        Offer.is_active == True,
        (Offer.end_date == None) | (Offer.end_date > now)
    ).all()
    
    # Activity logs
    logs = ActivityLog.query.filter(
        ActivityLog.timestamp >= start
    ).order_by(ActivityLog.timestamp.desc()).all()
    
    # Low stock products
    low_stock = Product.query.filter(Product.stock_quantity < 10).all()
    
    return {
        'period': period,
        'start_date': start.strftime('%Y-%m-%d'),
        'end_date': now.strftime('%Y-%m-%d'),
        'new_products_count': len(new_products),
        'new_offers_count': len(new_offers),
        'active_offers_count': len(active_offers),
        'low_stock_count': len(low_stock),
        'activity_count': len(logs),
        'new_products': [p.to_dict() for p in new_products],
        'new_offers': [o.to_dict() for o in new_offers],
        'active_offers': [o.to_dict() for o in active_offers],
        'low_stock_products': [p.to_dict() for p in low_stock],
        'activity_logs': [l.to_dict() for l in logs[:50]]
    }


def generate_ai_summary(report_data):
    """Generate AI summary of the report."""
    from ai.offer_generator import get_gemini_model
    
    model = get_gemini_model()
    
    if model is None:
        return _fallback_summary(report_data)
    
    prompt = f"""You are a retail business analyst. Summarize this store report in 3-4 sentences.

Period: {report_data['start_date']} to {report_data['end_date']} ({report_data['period']} report)
New Products: {report_data['new_products_count']}
New Offers: {report_data['new_offers_count']}
Active Offers: {report_data['active_offers_count']}
Low Stock Products: {report_data['low_stock_count']}
Activity Events: {report_data['activity_count']}

Write a concise, professional summary highlighting key trends and recommendations:"""
    
    try:
        response = model.generate_content(prompt)
        return response.text.strip()
    except Exception as e:
        return _fallback_summary(report_data)


def _fallback_summary(report_data):
    """Fallback summary without AI."""
    return (
        f"During the period {report_data['start_date']} to {report_data['end_date']}, "
        f"the store added {report_data['new_products_count']} new products and created "
        f"{report_data['new_offers_count']} promotional offers. Currently, there are "
        f"{report_data['active_offers_count']} active offers running. "
        f"{report_data['low_stock_count']} products need restocking attention."
    )


@reports_bp.route('/<period>', methods=['GET'])
@admin_required
def get_report(period):
    """Get report for specified period (daily/weekly/monthly)."""
    if period not in ['daily', 'weekly', 'monthly']:
        return jsonify({'error': 'Period must be daily, weekly, or monthly'}), 400
    
    report_data = get_report_data(period)
    summary = generate_ai_summary(report_data)
    report_data['ai_summary'] = summary
    
    log_activity('REPORT_GENERATED', f'{period} report generated')
    
    return jsonify(report_data), 200


@reports_bp.route('/<period>/export/<format>', methods=['GET'])
@admin_required
def export_report(period, format):
    """Export report as PDF or Excel."""
    if period not in ['daily', 'weekly', 'monthly']:
        return jsonify({'error': 'Invalid period'}), 400
    
    if format not in ['pdf', 'excel']:
        return jsonify({'error': 'Format must be pdf or excel'}), 400
    
    report_data = get_report_data(period)
    summary = generate_ai_summary(report_data)
    
    if format == 'pdf':
        return _export_pdf(report_data, summary)
    else:
        return _export_excel(report_data, summary)


def _export_pdf(report_data, summary):
    """Export report as PDF."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        elements.append(Paragraph(f"RetailMind AI - {report_data['period'].title()} Report", styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Period: {report_data['start_date']} to {report_data['end_date']}", styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Summary
        elements.append(Paragraph("AI Summary", styles['Heading2']))
        elements.append(Paragraph(summary, styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Stats
        elements.append(Paragraph("Key Metrics", styles['Heading2']))
        elements.append(Paragraph(f"New Products: {report_data['new_products_count']}", styles['Normal']))
        elements.append(Paragraph(f"New Offers: {report_data['new_offers_count']}", styles['Normal']))
        elements.append(Paragraph(f"Active Offers: {report_data['active_offers_count']}", styles['Normal']))
        elements.append(Paragraph(f"Low Stock Products: {report_data['low_stock_count']}", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"retailmind_{report_data['period']}_report.pdf"
        )
    except ImportError:
        return jsonify({'error': 'ReportLab not installed'}), 500


def _export_excel(report_data, summary):
    """Export report as Excel."""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_data['period'].title()} Report"
        
        # Header
        ws['A1'] = f"RetailMind AI - {report_data['period'].title()} Report"
        ws['A2'] = f"Period: {report_data['start_date']} to {report_data['end_date']}"
        ws['A4'] = 'AI Summary:'
        ws['A5'] = summary
        
        # Stats
        ws['A7'] = 'Metric'
        ws['B7'] = 'Value'
        ws['A8'] = 'New Products'
        ws['B8'] = report_data['new_products_count']
        ws['A9'] = 'New Offers'
        ws['B9'] = report_data['new_offers_count']
        ws['A10'] = 'Active Offers'
        ws['B10'] = report_data['active_offers_count']
        ws['A11'] = 'Low Stock Products'
        ws['B11'] = report_data['low_stock_count']
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"retailmind_{report_data['period']}_report.xlsx"
        )
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500
